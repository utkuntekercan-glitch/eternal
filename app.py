import hashlib
import io
import os
import re
import sqlite3
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st

try:
    import psycopg2
except Exception:
    psycopg2 = None

try:
    from psycopg2.extras import execute_batch
except Exception:
    execute_batch = None

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfgen import canvas
except Exception:
    A4 = None
    colors = None
    pdfmetrics = None
    TTFont = None
    canvas = None


st.set_page_config(page_title="Eternal Fire", layout="wide")

DB_PATH = Path("sales_reports.db")
DATABASE_URL = str(st.secrets.get("DATABASE_URL", os.getenv("DATABASE_URL", ""))).strip()
USE_POSTGRES = bool(DATABASE_URL)
APP_LOGO_URL = str(st.secrets.get("APP_LOGO_URL", os.getenv("APP_LOGO_URL", ""))).strip()
FREE_EXIT_EMAIL = "hakanerdgnn@gmail.com"
APP_USER = str(st.secrets.get("APP_USER", os.getenv("APP_USER", "admin"))).strip()
APP_PASSWORD = str(st.secrets.get("APP_PASSWORD", os.getenv("APP_PASSWORD", "1234"))).strip()


def stop_on_db_error(exc: Exception):
    st.session_state.pop("_sales_conn", None)
    st.session_state.pop("_sales_schema_ready", None)
    st.session_state.pop("_sales_dedupe_done", None)

    if USE_POSTGRES:
        st.error("Veritabani baglantisi kurulamadi. Streamlit secrets icindeki DATABASE_URL degerini kontrol edin.")
        st.info("PostgreSQL servisinin acik oldugundan, baglanti adresi/parola bilgilerinin guncel oldugundan ve gerekiyorsa SSL desteklediginden emin olun.")
    else:
        st.error("Yerel veritabani acilamadi. Uygulamanin sales_reports.db dosyasini olusturabildigini kontrol edin.")
    st.caption(f"Teknik hata: {exc.__class__.__name__}")
    detail = str(exc).strip()
    if detail:
        if DATABASE_URL:
            detail = detail.replace(DATABASE_URL, "[DATABASE_URL]")
        detail = re.sub(r"(postgres(?:ql)?://[^:\s]+:)[^@\s]+@", r"\1***@", detail, flags=re.IGNORECASE)
        detail = re.sub(r"(password\s*=\s*)[^\s]+", r"\1***", detail, flags=re.IGNORECASE)
        with st.expander("Teknik detay"):
            st.code(detail[:1000])
    st.stop()


def inject_styles():
    st.markdown(
        """
        <style>
        :root {
            --bg: #0f1218;
            --panel: #161b24;
            --panel-2: #1b2230;
            --text: #e8ebf2;
            --muted: #aeb8cb;
            --accent: #f0b429;
            --danger: #c7363e;
            --line: rgba(255,255,255,0.14);
        }
        .block-container {max-width: 1240px; padding-top: 0.8rem;}
        .stApp {
            background:
                radial-gradient(900px 320px at 4% -10%, rgba(240,180,41,0.18), transparent 55%),
                radial-gradient(880px 300px at 100% 0%, rgba(199,54,62,0.18), transparent 48%),
                var(--bg);
            color: var(--text);
        }
        .ef-shell {
            background: linear-gradient(165deg, #1a202c, #121722);
            border: 1px solid var(--line);
            border-radius: 14px;
            padding: 12px 14px;
            margin-bottom: 10px;
            box-shadow: 0 10px 24px rgba(0,0,0,0.25);
        }
        .ef-shell-title {
            font-size: 1.55rem;
            font-weight: 800;
            letter-spacing: 0.2px;
            color: #ffffff;
            margin: 0;
        }
        .ef-shell-sub {
            color: var(--muted);
            font-size: 0.92rem;
            margin: 2px 0 8px 0;
        }
        .ef-chip-row {
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
        }
        .ef-chip {
            border: 1px solid var(--line);
            border-radius: 999px;
            padding: 4px 10px;
            font-size: 0.78rem;
            color: #f7f8fb;
            background: rgba(255,255,255,0.05);
        }
        .stMetric {
            border: 1px solid var(--line);
            border-radius: 12px;
            padding: 8px;
            background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(255,255,255,0.02));
        }
        [data-testid="stMetricLabel"] {color: #cfd6e6 !important; font-weight: 700;}
        [data-testid="stMetricValue"] {color: #ffffff !important; font-weight: 800;}
        [data-testid="stMetricDelta"] {color: #8fd19e !important;}
        div[data-testid="stDataFrame"] {
            border: 1px solid var(--line);
            border-radius: 10px;
            background: var(--panel);
        }
        .stButton > button {
            border-radius: 10px;
            font-weight: 700;
            border: 1px solid rgba(240,180,41,0.45);
        }
        div[role="radiogroup"] {
            border: 1px solid var(--line);
            border-radius: 12px;
            padding: 6px;
            background: rgba(255,255,255,0.03);
        }
        div[role="radiogroup"] label {
            border: 1px solid rgba(255,255,255,0.16);
            border-radius: 9px;
            padding: 4px 10px;
            margin-right: 6px;
            background: rgba(255,255,255,0.03);
        }
        .stTextInput > div > div, .stSelectbox > div > div, .stDateInput > div > div {
            background: var(--panel-2);
            border-radius: 10px;
        }
        @media (max-width: 900px) {
            .block-container {
                padding-top: 0.3rem !important;
                padding-left: 0.7rem !important;
                padding-right: 0.7rem !important;
                max-width: 100% !important;
            }
            h1 {
                margin-top: 0.4rem !important;
                font-size: 1.7rem !important;
            }
            p {
                line-height: 1.35;
            }
            div[role="radiogroup"] {
                display: flex;
                flex-wrap: wrap;
                gap: 6px;
            }
            div[role="radiogroup"] label {
                margin-right: 0 !important;
                width: calc(50% - 4px);
                min-width: 120px;
            }
            .stButton > button {
                width: 100%;
            }
            div[data-testid="stDataFrame"] {
                font-size: 12px;
            }
            [data-testid="stMetricValue"] {
                font-size: 1.2rem !important;
            }
            [data-testid="stMetricLabel"] {
                font-size: 0.82rem !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


class DBConn:
    def __init__(self, driver: str, raw_conn):
        self.driver = driver
        self._conn = raw_conn

    def _sql(self, q: str) -> str:
        return q.replace("?", "%s") if self.driver == "postgres" else q

    def execute(self, q: str, params=()):
        try:
            cur = self._conn.cursor()
            cur.execute(self._sql(q), tuple(params))
            return cur
        except Exception:
            try:
                self._conn.rollback()
            except Exception:
                pass
            cur = self._conn.cursor()
            cur.execute(self._sql(q), tuple(params))
            return cur

    def executemany(self, q: str, seq):
        rows = [tuple(x) for x in seq]
        if not rows:
            return None
        try:
            cur = self._conn.cursor()
            if self.driver == "postgres" and execute_batch is not None:
                execute_batch(cur, self._sql(q), rows, page_size=500)
            else:
                cur.executemany(self._sql(q), rows)
            return cur
        except Exception:
            try:
                self._conn.rollback()
            except Exception:
                pass
            cur = self._conn.cursor()
            if self.driver == "postgres" and execute_batch is not None:
                execute_batch(cur, self._sql(q), rows, page_size=500)
            else:
                cur.executemany(self._sql(q), rows)
            return cur

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()


def get_conn() -> DBConn:
    if USE_POSTGRES:
        if psycopg2 is None:
            raise RuntimeError("Postgres icin psycopg2-binary gerekli.")
        kwargs = {
            "connect_timeout": 8,
            "application_name": "eternal-streamlit",
            "keepalives": 1,
            "keepalives_idle": 30,
            "keepalives_interval": 10,
            "keepalives_count": 3,
        }
        if "sslmode=" in DATABASE_URL:
            raw = psycopg2.connect(DATABASE_URL, **kwargs)
        else:
            raw = psycopg2.connect(DATABASE_URL, sslmode="require", **kwargs)
        raw.autocommit = False
        cur = raw.cursor()
        cur.execute("SET statement_timeout TO 12000")
        cur.execute("SET lock_timeout TO 3000")
        cur.close()
        return DBConn("postgres", raw)

    raw = sqlite3.connect(DB_PATH, check_same_thread=False)
    raw.execute("PRAGMA foreign_keys = ON;")
    return DBConn("sqlite", raw)


def init_db(conn: DBConn):
    id_col = "BIGSERIAL PRIMARY KEY" if conn.driver == "postgres" else "INTEGER PRIMARY KEY AUTOINCREMENT"
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS sales (
            id {id_col},
            week_label TEXT NOT NULL,
            order_date TEXT NOT NULL,
            ym TEXT,
            order_no TEXT,
            order_item_key TEXT,
            customer_email TEXT,
            is_free_exit INTEGER NOT NULL DEFAULT 0,
            is_returned INTEGER NOT NULL DEFAULT 0,
            sku TEXT NOT NULL,
            product_name TEXT NOT NULL,
            qty REAL NOT NULL,
            unit_price REAL NOT NULL,
            revenue REAL NOT NULL,
            order_total REAL NOT NULL DEFAULT 0,
            source_file TEXT NOT NULL,
            source_hash TEXT NOT NULL
        )
        """
    )
    # Safe incremental migrations.
    def ensure_sales_column(col_name: str, col_def: str):
        if conn.driver == "postgres":
            exists = df_query(
                conn,
                """
                SELECT 1
                FROM information_schema.columns
                WHERE table_schema='public' AND table_name='sales' AND column_name=?
                LIMIT 1
                """,
                (col_name,),
            )
            if exists.empty:
                conn.execute(f"ALTER TABLE sales ADD COLUMN {col_name} {col_def}")
        else:
            info = df_query(conn, "PRAGMA table_info(sales)")
            has_col = bool((info["name"] == col_name).any()) if (not info.empty and "name" in info.columns) else False
            if not has_col:
                conn.execute(f"ALTER TABLE sales ADD COLUMN {col_name} {col_def}")

    ensure_sales_column("free_exit_note", "TEXT")
    ensure_sales_column("is_returned", "INTEGER NOT NULL DEFAULT 0")
    ensure_sales_column("order_no", "TEXT")
    ensure_sales_column("order_item_key", "TEXT")
    ensure_sales_column("order_total", "REAL NOT NULL DEFAULT 0")
    ensure_sales_column("ym", "TEXT")
    if conn.driver == "postgres":
        conn.execute("UPDATE sales SET ym = SUBSTRING(order_date, 1, 7) WHERE COALESCE(ym, '') = '' AND COALESCE(order_date, '') <> ''")
    else:
        conn.execute("UPDATE sales SET ym = SUBSTR(order_date, 1, 7) WHERE COALESCE(ym, '') = '' AND COALESCE(order_date, '') <> ''")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            sku TEXT PRIMARY KEY,
            product_name TEXT NOT NULL,
            category TEXT NOT NULL DEFAULT 'Genel',
            unit_cost REAL NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS product_costs (
            sku TEXT PRIMARY KEY,
            unit_cost REAL NOT NULL DEFAULT 0,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sales_monthly_sku (
            ym TEXT NOT NULL,
            sku TEXT NOT NULL,
            product_name TEXT NOT NULL,
            qty REAL NOT NULL,
            revenue REAL NOT NULL,
            PRIMARY KEY (ym, sku)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS order_registry (
            order_no TEXT PRIMARY KEY,
            order_date TEXT,
            ym TEXT,
            customer_email TEXT,
            payment_status TEXT,
            is_free_exit INTEGER NOT NULL DEFAULT 0,
            is_returned INTEGER NOT NULL DEFAULT 0,
            source_file TEXT,
            source_hash TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS app_metrics (
            metric_key TEXT PRIMARY KEY,
            metric_value REAL NOT NULL DEFAULT 0,
            source_file TEXT,
            source_hash TEXT,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_order_date ON sales(order_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_week ON sales(week_label)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_sku ON sales(sku)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_free ON sales(is_free_exit)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_returned ON sales(is_returned)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_ym ON sales(ym)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_order_free ON sales(order_date, is_free_exit)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_order_returned ON sales(order_date, is_returned)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_order_no ON sales(order_no)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS uq_sales_order_item_key ON sales(order_item_key)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_monthly_ym ON sales_monthly_sku(ym)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_product_costs_sku ON product_costs(sku)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_order_registry_ym ON order_registry(ym)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_order_registry_source_hash ON order_registry(source_hash)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_app_metrics_source_hash ON app_metrics(source_hash)")
    conn.commit()


def get_ready_conn() -> DBConn:
    try:
        if "_sales_conn" not in st.session_state:
            st.session_state["_sales_conn"] = get_conn()
        conn = st.session_state["_sales_conn"]
        schema_key = f"{conn.driver}:clean-v4"
        if st.session_state.get("_sales_schema_ready") != schema_key:
            init_db(conn)
            st.session_state["_sales_schema_ready"] = schema_key
        return conn
    except Exception as exc:
        stop_on_db_error(exc)


def df_query(conn: DBConn, q: str, params=()):
    cur = conn.execute(q, params)
    rows = cur.fetchall()
    cols = [c[0] for c in cur.description] if cur.description else []
    return pd.DataFrame(rows, columns=cols)


def iter_chunks(values, chunk_size: int = 500):
    for start in range(0, len(values), chunk_size):
        yield values[start:start + chunk_size]


def select_months_for_order_nos(conn: DBConn, order_nos: list[str]) -> pd.DataFrame:
    frames = []
    for chunk in iter_chunks(order_nos):
        placeholders = ",".join(["?"] * len(chunk))
        frames.append(df_query(conn, f"SELECT DISTINCT ym FROM sales WHERE order_no IN ({placeholders})", tuple(chunk)))
    if not frames:
        return pd.DataFrame(columns=["ym"])
    return pd.concat(frames, ignore_index=True).drop_duplicates()


def delete_sales_by_order_nos(conn: DBConn, order_nos: list[str]):
    for chunk in iter_chunks(order_nos):
        placeholders = ",".join(["?"] * len(chunk))
        conn.execute(f"DELETE FROM sales WHERE order_no IN ({placeholders})", tuple(chunk))


def tr_money(x: float) -> str:
    return f"TL {format(float(x), ',.0f').replace(',', '.')}"


def normalize_num(x) -> float:
    if pd.isna(x):
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("TL", "").replace("tl", "")
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        parts = s.split(".")
        if len(parts) > 1 and all(len(p) == 3 for p in parts[1:]):
            s = "".join(parts)
    try:
        return float(s)
    except Exception:
        return 0.0


def is_valid_ym(ym: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}", ym.strip()))


def normalize_excel_date(value, fallback_iso: str) -> str:
    if value is None:
        return fallback_iso
    try:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, (int, float)):
            # Excel serial date (day 1 = 1899-12-31 with leap bug adjustment).
            serial = float(value)
            if serial > 1000:
                d = datetime(1899, 12, 30) + timedelta(days=serial)
                return d.strftime("%Y-%m-%d")
        s = str(value).strip()
        # Keep ISO dates deterministic (avoid day/month swap with dayfirst=True).
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?", s):
            return s[:10]
        # Turkish-style textual date (DD.MM.YYYY or DD/MM/YYYY).
        if re.fullmatch(r"\d{1,2}[./]\d{1,2}[./]\d{4}([ T]\d{2}:\d{2}(:\d{2})?)?", s):
            dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
        else:
            dt = pd.to_datetime(s, errors="coerce", dayfirst=False)
        if pd.isna(dt):
            return fallback_iso
        # Guard against bad parse (e.g., impossible historical/future years).
        if dt.year < 2020 or dt.year > (datetime.today().year + 1):
            return fallback_iso
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return fallback_iso


def normalize_header_text(s: str) -> str:
    tr_map = str.maketrans("Ã§ÄŸÄ±Ã¶ÅŸÃ¼Ã‡ÄÄ°Ã–ÅÃœ", "cgiosuCGIOSU")
    return str(s or "").translate(tr_map).strip().lower()


def normalize_text_safe(s: str) -> str:
    txt = str(s or "").strip().lower()
    # Fix common mojibake where utf-8 text is decoded as latin-1.
    try:
        txt = txt.encode("latin1").decode("utf-8")
    except Exception:
        pass
    txt = txt.replace("þ", "s").replace("ð", "g")
    txt = txt.replace("Ã¾", "s").replace("Ã°", "g")
    txt = txt.replace("ı", "i").replace("Ä±", "i")
    txt = txt.translate(str.maketrans("çğıöşüÇĞİÖŞÜ", "cgiosuCGIOSU"))
    txt = txt.replace("\u0131", "i")
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    return txt


def find_order_no_col(ws) -> int | None:
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except Exception:
        return None
    for i, v in enumerate(header_row):
        h = normalize_text_safe(v)
        if ("siparis" in h or "order" in h) and ("no" in h or "numara" in h):
            return i
    return None


def build_order_item_key(
    order_no: str,
    order_date: str,
    sku: str,
    qty: float,
    unit_price: float,
    customer_email: str,
    source_hash: str,
    excel_row_no: int,
) -> str:
    # Prefer stable business key when order number exists.
    if order_no:
        base = f"ord:{order_no}|{order_date}|{sku}|{qty:.6f}|{unit_price:.6f}|{customer_email}"
    else:
        # Fallback keeps each source row unique while still deduping re-import of same file.
        base = f"src:{source_hash}|row:{excel_row_no}|{order_date}|{sku}|{qty:.6f}|{unit_price:.6f}|{customer_email}"
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


def is_paid_status(v) -> bool:
    t = normalize_text_safe(v)
    t = " ".join(t.split())
    return ("odendi" in t) or ("iade edildi" in t)


def is_returned_status(v) -> bool:
    t = normalize_text_safe(v)
    t = " ".join(t.split())
    return "iade edildi" in t


def row_value(row, idx: int, default=None):
    if idx is None or idx < 0 or idx >= len(row):
        return default
    return row[idx].value


def get_header_map(ws) -> dict[str, int]:
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except Exception:
        return {}
    headers = {}
    for i, value in enumerate(header_row):
        key = " ".join(normalize_text_safe(value).split())
        if key and key not in headers:
            headers[key] = i
    return headers


def find_header_idx(headers: dict[str, int], required: tuple[str, ...], forbidden: tuple[str, ...] = ()) -> int | None:
    for header, idx in headers.items():
        if all(token in header for token in required) and not any(token in header for token in forbidden):
            return idx
    return None


def first_header_idx(headers: dict[str, int], *names: str) -> int | None:
    for name in names:
        idx = headers.get(name)
        if idx is not None:
            return idx
    return None


def header_idx_or(default: int, idx: int | None) -> int:
    return default if idx is None else idx


def make_summary_sku(product_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", normalize_text_safe(product_name)).strip("-").upper()
    if not slug:
        return ""
    return f"SUMMARY-{slug[:64]}"


def parse_order_detail_rows(ws, week_label: str) -> list[dict]:
    headers = get_header_map(ws)
    order_no_idx = find_order_no_col(ws)
    email_idx = header_idx_or(4, first_header_idx(headers, "e-posta", "email"))
    status_idx = header_idx_or(6, find_header_idx(headers, ("siparis", "odeme", "durumu")))
    order_date_idx = header_idx_or(7, find_header_idx(headers, ("siparis", "tarihi")))
    qty_idx = header_idx_or(17, first_header_idx(headers, "urun sayisi"))
    product_idx = header_idx_or(18, first_header_idx(headers, "urun adi", "urun"))
    unit_price_idx = header_idx_or(20, first_header_idx(headers, "urun satis fiyati", "birim fiyat"))
    order_total_idx = header_idx_or(15, first_header_idx(headers, "toplam"))
    sku_idx = header_idx_or(24, first_header_idx(headers, "urun sku", "sku"))
    order_context = {}
    rows = []
    for excel_row_no, r in enumerate(ws.iter_rows(min_row=2), start=2):
        raw_order_no = row_value(r, order_no_idx)
        if raw_order_no is None:
            order_no = ""
        elif isinstance(raw_order_no, float) and raw_order_no.is_integer():
            order_no = str(int(raw_order_no))
        else:
            order_no = str(raw_order_no).strip()

        context = order_context.get(order_no, {}) if order_no else {}
        customer_email = "" if row_value(r, email_idx) is None else str(row_value(r, email_idx)).strip().lower()
        if not customer_email:
            customer_email = str(context.get("customer_email", ""))
        status_value = row_value(r, status_idx)
        if (status_value is None or str(status_value).strip() == "") and context:
            status_value = context.get("status_value")
        is_returned = 1 if is_returned_status(status_value) else 0
        is_free_exit = 1 if customer_email == FREE_EXIT_EMAIL else 0
        if (not is_paid_status(status_value)) and (is_free_exit == 0):
            continue
        order_date = normalize_excel_date(row_value(r, order_date_idx), "")
        if not order_date:
            order_date = str(context.get("order_date", ""))
        if not order_date:
            continue
        if order_no:
            order_context[order_no] = {
                "customer_email": customer_email,
                "status_value": status_value,
                "order_date": order_date,
            }
        qty = normalize_num(row_value(r, qty_idx))
        product_name = "" if row_value(r, product_idx) is None else str(row_value(r, product_idx)).strip()
        unit_price = normalize_num(row_value(r, unit_price_idx))
        order_total = normalize_num(row_value(r, order_total_idx))
        sku = "" if row_value(r, sku_idx) is None else str(row_value(r, sku_idx)).strip()
        if sku and product_name and qty > 0:
            rows.append(
                {
                    "week_label": week_label,
                    "order_date": order_date,
                    "ym": order_date[:7],
                    "order_no": order_no,
                    "excel_row_no": int(excel_row_no),
                    "customer_email": customer_email,
                    "is_free_exit": is_free_exit,
                    "is_returned": is_returned,
                    "sku": sku,
                    "product_name": product_name,
                    "qty": float(qty),
                    "unit_price": float(unit_price),
                    "revenue": float(qty) * float(unit_price),
                    "order_total": float(order_total),
                }
            )
    return rows


def parse_order_registry_rows(ws) -> list[dict]:
    headers = get_header_map(ws)
    if not headers:
        return []

    order_no_idx = find_order_no_col(ws)
    if order_no_idx is None:
        return []

    email_idx = header_idx_or(4, first_header_idx(headers, "e-posta", "email"))
    status_idx = header_idx_or(6, find_header_idx(headers, ("siparis", "odeme", "durumu")))
    order_date_idx = header_idx_or(7, find_header_idx(headers, ("siparis", "tarihi")))

    orders = {}
    order_context = {}
    for r in ws.iter_rows(min_row=2):
        raw_order_no = row_value(r, order_no_idx)
        if raw_order_no is None:
            continue
        if isinstance(raw_order_no, float) and raw_order_no.is_integer():
            order_no = str(int(raw_order_no))
        else:
            order_no = str(raw_order_no).strip()
        if not order_no:
            continue

        context = order_context.get(order_no, {})
        customer_email = "" if row_value(r, email_idx) is None else str(row_value(r, email_idx)).strip().lower()
        if not customer_email:
            customer_email = str(context.get("customer_email", ""))
        status_value = row_value(r, status_idx)
        if (status_value is None or str(status_value).strip() == "") and context:
            status_value = context.get("status_value")
        payment_status = "" if status_value is None else str(status_value).strip()
        order_date = normalize_excel_date(row_value(r, order_date_idx), "")
        if not order_date:
            order_date = str(context.get("order_date", ""))

        if order_date or payment_status or customer_email:
            order_context[order_no] = {
                "customer_email": customer_email,
                "status_value": status_value,
                "order_date": order_date,
            }
        if order_no not in orders:
            orders[order_no] = {
                "order_no": order_no,
                "order_date": order_date,
                "ym": order_date[:7] if order_date else "",
                "customer_email": customer_email,
                "payment_status": payment_status,
                "is_free_exit": 1 if customer_email == FREE_EXIT_EMAIL else 0,
                "is_returned": 1 if is_returned_status(status_value) else 0,
            }
        else:
            current = orders[order_no]
            if not current.get("order_date") and order_date:
                current["order_date"] = order_date
                current["ym"] = order_date[:7]
            if not current.get("customer_email") and customer_email:
                current["customer_email"] = customer_email
                current["is_free_exit"] = 1 if customer_email == FREE_EXIT_EMAIL else 0
            if not current.get("payment_status") and payment_status:
                current["payment_status"] = payment_status
            if is_returned_status(status_value):
                current["is_returned"] = 1

    return list(orders.values())


def parse_product_summary_rows(ws, week_label: str, order_date_iso: str) -> list[dict]:
    headers = get_header_map(ws)
    product_idx = headers.get("urun")
    type_idx = find_header_idx(headers, ("siparis", "kalem", "tipi"))
    qty_idx = find_header_idx(headers, ("sepetteki", "toplam", "urun", "sayisi"))
    revenue_idx = find_header_idx(headers, ("net", "satis", "tutari"), ("doviz",))
    avg_price_idx = find_header_idx(headers, ("ortalama", "satis", "fiyati"), ("doviz",))

    if product_idx is None or qty_idx is None or revenue_idx is None:
        return []

    rows = []
    for excel_row_no, r in enumerate(ws.iter_rows(min_row=2), start=2):
        product_name = "" if row_value(r, product_idx) is None else str(row_value(r, product_idx)).strip()
        item_type = normalize_text_safe(row_value(r, type_idx)) if type_idx is not None else "urun"
        if not product_name or (item_type and item_type != "urun"):
            continue

        qty = normalize_num(row_value(r, qty_idx))
        revenue = normalize_num(row_value(r, revenue_idx))
        unit_price = normalize_num(row_value(r, avg_price_idx)) if avg_price_idx is not None else 0.0
        if unit_price <= 0 and qty > 0:
            unit_price = revenue / qty

        sku = make_summary_sku(product_name)
        if not sku or qty <= 0:
            continue

        order_date = normalize_excel_date(order_date_iso, datetime.today().date().isoformat())
        rows.append(
            {
                "week_label": week_label,
                "order_date": order_date,
                "ym": order_date[:7],
                "order_no": f"summary:{order_date}:{sku}",
                "excel_row_no": int(excel_row_no),
                "customer_email": "",
                "is_free_exit": 0,
                "is_returned": 0,
                "sku": sku,
                "product_name": product_name,
                "qty": float(qty),
                "unit_price": float(unit_price),
                "revenue": float(revenue),
                "order_total": float(revenue),
            }
        )
    return rows


def parse_product_summary_metrics(ws) -> dict[str, float]:
    headers = get_header_map(ws)
    product_idx = headers.get("urun")
    type_idx = find_header_idx(headers, ("siparis", "kalem", "tipi"))
    refund_idx = find_header_idx(headers, ("iade", "tutari"), ("doviz",))
    net_sales_idx = find_header_idx(headers, ("net", "satis", "tutari"), ("doviz",))
    gross_sales_idx = find_header_idx(headers, ("brut", "satis", "tutari"), ("doviz",))
    sales_count_idx = first_header_idx(headers, "satislar")

    if product_idx is None or refund_idx is None:
        return {}

    metrics = {
        "ikas_summary_refund_amount": 0.0,
        "ikas_summary_net_sales": 0.0,
        "ikas_summary_gross_sales": 0.0,
        "ikas_summary_sales_count": 0.0,
    }
    found_product_row = False
    for r in ws.iter_rows(min_row=2):
        product_name = "" if row_value(r, product_idx) is None else str(row_value(r, product_idx)).strip()
        item_type = normalize_text_safe(row_value(r, type_idx)) if type_idx is not None else "urun"
        if not product_name or (item_type and item_type != "urun"):
            continue
        found_product_row = True
        metrics["ikas_summary_refund_amount"] += normalize_num(row_value(r, refund_idx))
        if net_sales_idx is not None:
            metrics["ikas_summary_net_sales"] += normalize_num(row_value(r, net_sales_idx))
        if gross_sales_idx is not None:
            metrics["ikas_summary_gross_sales"] += normalize_num(row_value(r, gross_sales_idx))
        if sales_count_idx is not None:
            metrics["ikas_summary_sales_count"] += normalize_num(row_value(r, sales_count_idx))
    return metrics if found_product_row else {}


def parse_uploaded_excel(file_bytes: bytes, week_label: str, order_date_iso: str) -> pd.DataFrame:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    order_registry_rows = parse_order_registry_rows(ws)
    rows = parse_order_detail_rows(ws, week_label)
    summary_metrics = parse_product_summary_metrics(ws)
    if not rows and not summary_metrics:
        rows = parse_product_summary_rows(ws, week_label, order_date_iso)
    df = pd.DataFrame(rows)
    df.attrs["order_registry_rows"] = order_registry_rows
    df.attrs["summary_metrics"] = summary_metrics
    return df


def upsert_products_from_rows(conn: DBConn, rows_df: pd.DataFrame):
    if rows_df.empty:
        return
    uniq = rows_df[["sku", "product_name"]].dropna().drop_duplicates()
    now = datetime.now().isoformat(timespec="seconds")
    payload = [(str(r["sku"]), str(r["product_name"]), now) for _, r in uniq.iterrows()]
    conn.executemany(
        """
        INSERT INTO products(sku, product_name, updated_at)
        VALUES(?,?,?)
        ON CONFLICT(sku) DO UPDATE SET
            product_name=excluded.product_name,
            updated_at=excluded.updated_at
        """,
        payload,
    )
    conn.commit()


def upsert_order_registry(conn: DBConn, orders: list[dict], source_file: str, source_hash: str):
    if not orders:
        return
    now = datetime.now().isoformat(timespec="seconds")
    payload = [
        (
            str(o.get("order_no", "") or "").strip(),
            str(o.get("order_date", "") or "").strip(),
            str(o.get("ym", "") or "").strip(),
            str(o.get("customer_email", "") or "").strip(),
            str(o.get("payment_status", "") or "").strip(),
            int(o.get("is_free_exit", 0) or 0),
            int(o.get("is_returned", 0) or 0),
            source_file,
            source_hash,
            now,
        )
        for o in orders
        if str(o.get("order_no", "") or "").strip()
    ]
    conn.executemany(
        """
        INSERT INTO order_registry(order_no, order_date, ym, customer_email, payment_status, is_free_exit, is_returned, source_file, source_hash, updated_at)
        VALUES(?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(order_no) DO UPDATE SET
            order_date=excluded.order_date,
            ym=excluded.ym,
            customer_email=excluded.customer_email,
            payment_status=excluded.payment_status,
            is_free_exit=excluded.is_free_exit,
            is_returned=excluded.is_returned,
            source_file=excluded.source_file,
            source_hash=excluded.source_hash,
            updated_at=excluded.updated_at
        """,
        payload,
    )
    conn.commit()


def upsert_app_metrics(conn: DBConn, metrics: dict[str, float], source_file: str, source_hash: str):
    if not metrics:
        return
    now = datetime.now().isoformat(timespec="seconds")
    payload = [
        (str(key), float(value or 0), source_file, source_hash, now)
        for key, value in metrics.items()
    ]
    conn.executemany(
        """
        INSERT INTO app_metrics(metric_key, metric_value, source_file, source_hash, updated_at)
        VALUES(?,?,?,?,?)
        ON CONFLICT(metric_key) DO UPDATE SET
            metric_value=excluded.metric_value,
            source_file=excluded.source_file,
            source_hash=excluded.source_hash,
            updated_at=excluded.updated_at
        """,
        payload,
    )
    conn.commit()


def upsert_product_costs(conn: DBConn, rows):
    now = datetime.now().isoformat(timespec="seconds")
    payload = [(r[0], float(r[1]), now) for r in rows]
    conn.executemany(
        """
        INSERT INTO product_costs(sku, unit_cost, updated_at)
        VALUES(?,?,?)
        ON CONFLICT(sku) DO UPDATE SET
            unit_cost=excluded.unit_cost,
            updated_at=excluded.updated_at
        """,
        payload,
    )
    conn.commit()


def refresh_monthly_summary_for_month(conn: DBConn, ym: str):
    conn.execute("DELETE FROM sales_monthly_sku WHERE ym=?", (ym,))
    conn.execute(
        """
        INSERT INTO sales_monthly_sku(ym, sku, product_name, qty, revenue)
        SELECT
            ? AS ym,
            sku,
            MAX(product_name) AS product_name,
            SUM(qty) AS qty,
            SUM(revenue) AS revenue
        FROM sales
        WHERE ym = ?
          AND COALESCE(is_free_exit, 0) = 0
          AND COALESCE(is_returned, 0) = 0
        GROUP BY sku
        """,
        (ym, ym),
    )
    conn.commit()


def refresh_monthly_summary_all(conn: DBConn):
    conn.execute("DELETE FROM sales_monthly_sku")
    conn.execute(
        """
        INSERT INTO sales_monthly_sku(ym, sku, product_name, qty, revenue)
        SELECT
            COALESCE(ym, SUBSTR(order_date, 1, 7)) AS ym,
            sku,
            MAX(product_name) AS product_name,
            SUM(qty) AS qty,
            SUM(revenue) AS revenue
        FROM sales
        WHERE COALESCE(is_free_exit, 0) = 0
          AND COALESCE(is_returned, 0) = 0
          AND COALESCE(ym, '') <> ''
        GROUP BY COALESCE(ym, SUBSTR(order_date, 1, 7)), sku
        """
    )
    conn.commit()


def dedupe_sales_by_order(conn: DBConn):
    conn.execute(
        """
        DELETE FROM sales
        WHERE id IN (
            SELECT s1.id
            FROM sales s1
            JOIN (
                SELECT
                    order_no,
                    sku,
                    qty,
                    unit_price,
                    customer_email,
                    is_free_exit,
                    is_returned,
                    MAX(id) AS keep_id,
                    COUNT(*) AS c
                FROM sales
                WHERE COALESCE(order_no, '') <> ''
                GROUP BY order_no, sku, qty, unit_price, customer_email, is_free_exit, is_returned
                HAVING COUNT(*) > 1
            ) d
              ON s1.order_no = d.order_no
             AND s1.sku = d.sku
             AND s1.qty = d.qty
             AND s1.unit_price = d.unit_price
             AND COALESCE(s1.customer_email, '') = COALESCE(d.customer_email, '')
             AND COALESCE(s1.is_free_exit, 0) = COALESCE(d.is_free_exit, 0)
             AND COALESCE(s1.is_returned, 0) = COALESCE(d.is_returned, 0)
            WHERE s1.id <> d.keep_id
        )
        """
    )
    conn.commit()


@st.cache_data(ttl=300, show_spinner=False)
def get_uploads(_conn: DBConn) -> pd.DataFrame:
    latest = df_query(
        _conn,
        """
        SELECT week_label, source_file, source_hash, id AS max_id
        FROM sales
        ORDER BY id DESC
        LIMIT 1
        """,
    )
    if latest.empty:
        return pd.DataFrame(columns=["week_label", "source_file", "source_hash", "row_count", "max_id"])
    source_hash = str(latest.iloc[0]["source_hash"] or "").strip()
    row_count = 0
    if source_hash:
        count_df = df_query(
            _conn,
            """
            SELECT COUNT(*) AS row_count
            FROM sales
            WHERE source_hash=?
            """,
            (source_hash,),
        )
        row_count = int(count_df.iloc[0]["row_count"] or 0) if not count_df.empty else 0
    return pd.DataFrame(
        [
            {
                "week_label": latest.iloc[0]["week_label"],
                "source_file": latest.iloc[0]["source_file"],
                "source_hash": source_hash,
                "row_count": row_count,
                "max_id": latest.iloc[0]["max_id"],
            }
        ]
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_recent_sales(_conn: DBConn, limit_n: int = 300) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT id, ym, order_date, week_label, sku, product_name, qty, revenue
        FROM sales
        ORDER BY id DESC
        LIMIT ?
        """,
        (int(limit_n),),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_metrics(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            COALESCE(SUM(m.qty), 0) AS total_qty,
            COALESCE(SUM(m.revenue), 0) AS total_revenue,
            COALESCE(SUM(m.qty * COALESCE(c.unit_cost, p.unit_cost, 0)), 0) AS total_cost,
            COALESCE(SUM(m.revenue - (m.qty * COALESCE(c.unit_cost, p.unit_cost, 0))), 0) AS total_profit
        FROM sales_monthly_sku m
        LEFT JOIN products p ON p.sku = m.sku
        LEFT JOIN product_costs c ON c.sku = m.sku
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_order_total(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        WITH order_totals AS (
            SELECT
                order_no,
                MAX(CASE WHEN COALESCE(order_total, 0) > 0 THEN order_total ELSE 0 END) AS order_total,
                SUM(revenue) AS revenue,
                MAX(COALESCE(is_free_exit, 0)) AS is_free_exit,
                MAX(COALESCE(is_returned, 0)) AS is_returned
            FROM sales
            WHERE COALESCE(order_no, '') <> ''
            GROUP BY order_no
        ),
        summary_metrics AS (
            SELECT
                MAX(CASE WHEN metric_key = 'ikas_summary_refund_amount' THEN metric_value ELSE 0 END) AS summary_refund_amount
            FROM app_metrics
        )
        SELECT
            COALESCE(
                NULLIF(SUM(
                    CASE
                        WHEN COALESCE(is_free_exit, 0) = 0
                         AND COALESCE(order_total, 0) > 0
                        THEN order_total
                        ELSE 0
                    END
                ), 0),
                SUM(CASE WHEN COALESCE(is_free_exit, 0) = 0 THEN revenue ELSE 0 END),
                0
            ) AS gross_order_revenue,
            COALESCE(
                NULLIF(SUM(
                    CASE
                        WHEN COALESCE(is_free_exit, 0) = 0
                         AND COALESCE(is_returned, 0) = 0
                         AND COALESCE(order_total, 0) > 0
                        THEN order_total
                        ELSE 0
                    END
                ), 0),
                SUM(CASE WHEN COALESCE(is_free_exit, 0) = 0 AND COALESCE(is_returned, 0) = 0 THEN revenue ELSE 0 END),
                0
            ) AS net_order_revenue,
            COALESCE(
                NULLIF((SELECT summary_refund_amount FROM summary_metrics), 0),
                SUM(
                    CASE
                        WHEN COALESCE(is_free_exit, 0) = 0
                         AND COALESCE(is_returned, 0) = 1
                        THEN CASE WHEN COALESCE(order_total, 0) > 0 THEN order_total ELSE revenue END
                        ELSE 0
                    END
                ),
                0
            ) AS returned_order_revenue
        FROM order_totals
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_order_rows(_conn: DBConn) -> pd.DataFrame:
    registry_count = df_query(
        _conn,
        """
        SELECT COUNT(DISTINCT order_no) AS order_rows
        FROM order_registry
        WHERE COALESCE(order_no, '') <> ''
        """,
    )
    if not registry_count.empty and int(registry_count.iloc[0]["order_rows"] or 0) > 0:
        return registry_count
    return df_query(
        _conn,
        """
        SELECT COUNT(DISTINCT order_no) AS order_rows
        FROM sales
        WHERE COALESCE(order_no, '') <> ''
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_date_range(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            MIN(order_date) AS min_date,
            MAX(order_date) AS max_date
        FROM sales
        WHERE COALESCE(is_free_exit, 0) = 0
          AND COALESCE(is_returned, 0) = 0
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_available_months(_conn: DBConn, free_only: int = 0) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT DISTINCT ym
        FROM sales
        WHERE COALESCE(ym, '') <> ''
          AND COALESCE(is_free_exit, 0) = ?
          AND COALESCE(is_returned, 0) = 0
        ORDER BY ym DESC
        """,
        (int(free_only),),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_top_products(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT sku AS "Stok Kodu", MAX(product_name) AS "Urun", SUM(qty) AS "Adet"
        FROM sales_monthly_sku
        GROUP BY sku
        ORDER BY SUM(qty) DESC
        LIMIT 10
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_dashboard_categories(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            COALESCE(p.category, 'Genel') AS "Kategori",
            COALESCE(
                SUM(CASE WHEN COALESCE(s.order_total, 0) > 0 THEN s.order_total ELSE s.revenue END),
                0
            ) AS "Ciro",
            COALESCE(
                SUM(
                    (CASE WHEN COALESCE(s.order_total, 0) > 0 THEN s.order_total ELSE s.revenue END)
                    - (s.qty * COALESCE(c.unit_cost, p.unit_cost, 0))
                ),
                0
            ) AS "Kar"
        FROM sales s
        LEFT JOIN products p ON p.sku = s.sku
        LEFT JOIN product_costs c ON c.sku = s.sku
        WHERE COALESCE(s.is_free_exit, 0) = 0
          AND COALESCE(s.is_returned, 0) = 0
        GROUP BY COALESCE(p.category, 'Genel')
        ORDER BY "Ciro" DESC
        """
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_month_products(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            s.sku AS "Stok Kodu",
            MAX(s.product_name) AS "Urun",
            COALESCE(SUM(s.qty), 0) AS "Adet",
            COALESCE(SUM(s.revenue), 0) AS "Ciro",
            COALESCE(SUM(s.qty * COALESCE(c.unit_cost, p.unit_cost, 0)), 0) AS "Maliyet",
            COALESCE(SUM(s.revenue - (s.qty * COALESCE(c.unit_cost, p.unit_cost, 0))), 0) AS "Kar",
            MAX(s.order_date) AS "Son Satis Tarihi"
        FROM sales s
        LEFT JOIN products p ON p.sku = s.sku
        LEFT JOIN product_costs c ON c.sku = s.sku
        WHERE s.ym = ?
          AND COALESCE(s.is_free_exit, 0) = 0
          AND COALESCE(s.is_returned, 0) = 0
        GROUP BY s.sku
        ORDER BY "Ciro" DESC
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_month_totals(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            COALESCE(SUM(s.qty), 0) AS total_qty,
            COALESCE(SUM(s.revenue), 0) AS total_revenue,
            COALESCE(SUM(s.qty * COALESCE(c.unit_cost, p.unit_cost, 0)), 0) AS total_cost,
            COALESCE(SUM(s.revenue - (s.qty * COALESCE(c.unit_cost, p.unit_cost, 0))), 0) AS total_profit
        FROM sales s
        LEFT JOIN products p ON p.sku = s.sku
        LEFT JOIN product_costs c ON c.sku = s.sku
        WHERE s.ym = ?
          AND COALESCE(s.is_free_exit, 0) = 0
          AND COALESCE(s.is_returned, 0) = 0
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_month_order_total(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        WITH order_totals AS (
            SELECT
                order_no,
                MAX(CASE WHEN COALESCE(order_total, 0) > 0 THEN order_total ELSE 0 END) AS order_total,
                SUM(revenue) AS revenue,
                MAX(COALESCE(is_free_exit, 0)) AS is_free_exit,
                MAX(COALESCE(is_returned, 0)) AS is_returned
            FROM sales
            WHERE ym = ?
              AND COALESCE(order_no, '') <> ''
            GROUP BY order_no
        )
        SELECT
            COALESCE(
                NULLIF(SUM(CASE WHEN COALESCE(order_total, 0) > 0 THEN order_total ELSE 0 END), 0),
                SUM(revenue),
                0
            ) AS total_order_revenue
        FROM order_totals
        WHERE COALESCE(is_free_exit, 0) = 0
          AND COALESCE(is_returned, 0) = 0
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_month_sku_details(_conn: DBConn, ym: str, sku: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            week_label AS "Hafta",
            order_date AS "Tarih",
            sku AS "Stok Kodu",
            product_name AS "Urun",
            qty AS "Adet",
            unit_price AS "Birim Fiyat",
            revenue AS "Ciro"
        FROM sales
        WHERE ym = ?
          AND COALESCE(is_free_exit, 0) = 0
          AND COALESCE(is_returned, 0) = 0
          AND sku = ?
        ORDER BY order_date, week_label
        """,
        (ym, sku),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_month_free_exit_rows(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            week_label AS "Hafta",
            order_date AS "Tarih",
            customer_email AS "E-Posta",
            sku AS "Stok Kodu",
            product_name AS "Urun",
            qty AS "Adet",
            unit_price AS "Birim Fiyat",
            revenue AS "Bedelsiz Tutar"
        FROM sales
        WHERE ym = ?
          AND COALESCE(is_free_exit, 0) = 1
        ORDER BY order_date, week_label
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_free_exit_manage_rows(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            id AS "ID",
            order_date AS "Tarih",
            customer_email AS "E-Posta",
            sku AS "Stok Kodu",
            product_name AS "Urun",
            qty AS "Adet",
            unit_price AS "Birim Fiyat",
            revenue AS "_BedelsizTutar",
            COALESCE(free_exit_note, '') AS "Aciklama"
        FROM sales
        WHERE ym = ?
          AND COALESCE(is_free_exit, 0) = 1
        ORDER BY order_date DESC, id DESC
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_returned_months(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT DISTINCT ym
        FROM sales
        WHERE COALESCE(ym, '') <> ''
          AND COALESCE(is_returned, 0) = 1
        ORDER BY ym DESC
        """,
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_returned_rows(_conn: DBConn, ym: str) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            order_date AS "Tarih",
            week_label AS "Hafta",
            customer_email AS "E-Posta",
            sku AS "Stok Kodu",
            product_name AS "Urun",
            qty AS "Adet",
            unit_price AS "Birim Fiyat",
            revenue AS "Iade Tutar"
        FROM sales
        WHERE ym = ?
          AND COALESCE(is_returned, 0) = 1
        ORDER BY order_date DESC, id DESC
        """,
        (ym,),
    )


@st.cache_data(ttl=300, show_spinner=False)
def get_products_master(_conn: DBConn) -> pd.DataFrame:
    return df_query(
        _conn,
        """
        SELECT
            p.sku,
            p.product_name,
            p.category,
            COALESCE(c.unit_cost, p.unit_cost, 0) AS unit_cost,
            p.active
        FROM products p
        LEFT JOIN product_costs c ON c.sku = p.sku
        ORDER BY product_name
        """,
    )


def _get_pdf_fonts() -> tuple[str, str]:
    if pdfmetrics is None or TTFont is None:
        return ("Helvetica", "Helvetica-Bold")

    regular_name = "EFDejaVuSans"
    bold_name = "EFDejaVuSansBold"
    registered = set(pdfmetrics.getRegisteredFontNames())
    if regular_name in registered and bold_name in registered:
        return (regular_name, bold_name)

    candidates = [
        (
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        ),
        (
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/arialbd.ttf",
        ),
    ]
    for reg_path, bold_path in candidates:
        try:
            if regular_name not in registered:
                pdfmetrics.registerFont(TTFont(regular_name, reg_path))
            if bold_name not in registered:
                pdfmetrics.registerFont(TTFont(bold_name, bold_path))
            return (regular_name, bold_name)
        except Exception:
            continue
    return ("Helvetica", "Helvetica-Bold")


def build_month_pdf(ym: str, totals_df: pd.DataFrame, order_total_df: pd.DataFrame, prod_df: pd.DataFrame) -> bytes | None:
    if canvas is None or A4 is None or colors is None:
        return None

    font_regular, font_bold = _get_pdf_fonts()
    total_qty = float(totals_df.iloc[0]["total_qty"] or 0) if not totals_df.empty else 0.0
    total_rev = float(order_total_df.iloc[0]["total_order_revenue"] or 0) if not order_total_df.empty else 0.0
    total_cost = float(totals_df.iloc[0]["total_cost"] or 0) if not totals_df.empty else 0.0
    total_profit = total_rev - total_cost

    def money(v: float) -> str:
        return f"â‚º{format(float(v), ',.2f').replace(',', '.')}"

    out = io.BytesIO()
    pdf = canvas.Canvas(out, pagesize=A4)
    w, h = A4
    margin = 36

    def draw_header(page_no: int):
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.rect(0, h - 84, w, 84, stroke=0, fill=1)
        pdf.setFillColor(colors.HexColor("#f59e0b"))
        pdf.rect(0, h - 84, 7, 84, stroke=0, fill=1)

        pdf.setFillColor(colors.white)
        pdf.setFont(font_bold, 17)
        pdf.drawString(margin, h - 34, "Eternal Fire")
        pdf.setFont(font_regular, 11)
        pdf.drawString(margin, h - 52, "AylÄ±k SatÄ±ÅŸ ve KarlÄ±lÄ±k Raporu")
        pdf.setFont(font_regular, 9)
        pdf.drawRightString(w - margin, h - 36, f"Rapor AyÄ±: {ym}")
        pdf.drawRightString(
            w - margin,
            h - 52,
            f"OluÅŸturma: {datetime.now(ZoneInfo('Europe/Istanbul')).strftime('%d.%m.%Y %H:%M')}",
        )
        pdf.drawRightString(w - margin, h - 68, f"Sayfa {page_no}")

        card_top = h - 98
        card_h = 40
        card_w = (w - (margin * 2) - 18) / 4
        card_bg = colors.HexColor("#f3f4f6")
        metrics = [
            ("Toplam Adet", format(total_qty, ",.0f").replace(",", ".")),
            ("Toplam Ciro", money(total_rev)),
            ("Toplam Maliyet", money(total_cost)),
            ("Net Kar", money(total_profit)),
        ]
        for i, (title, value) in enumerate(metrics):
            x = margin + i * (card_w + 6)
            pdf.setFillColor(card_bg)
            pdf.roundRect(x, card_top - card_h, card_w, card_h, 6, stroke=0, fill=1)
            pdf.setFillColor(colors.HexColor("#374151"))
            pdf.setFont(font_regular, 8)
            pdf.drawString(x + 8, card_top - 14, title)
            pdf.setFillColor(colors.HexColor("#111827"))
            pdf.setFont(font_bold, 11)
            pdf.drawString(x + 8, card_top - 30, value)

        table_top = card_top - card_h - 14
        table_w = w - (margin * 2)
        col_w = {
            "sku": 88,
            "urun": 162,
            "adet": 55,
            "ciro": 70,
            "maliyet": 70,
            "kar": max(60, table_w - (88 + 162 + 55 + 70 + 70)),
        }
        col_left = {}
        col_right = {}
        x_cur = margin
        for key in ["sku", "urun", "adet", "ciro", "maliyet", "kar"]:
            col_left[key] = x_cur
            x_cur += col_w[key]
            col_right[key] = x_cur
        pdf.setFillColor(colors.HexColor("#1f2937"))
        pdf.rect(margin, table_top - 18, table_w, 18, stroke=0, fill=1)
        pdf.setFillColor(colors.white)
        pdf.setFont(font_bold, 8.5)
        pdf.drawString(col_left["sku"] + 4, table_top - 12, "SKU")
        pdf.drawString(col_left["urun"] + 4, table_top - 12, "ÃœrÃ¼n")
        pdf.drawRightString(col_right["adet"] - 4, table_top - 12, "Adet")
        pdf.drawRightString(col_right["ciro"] - 4, table_top - 12, "Ciro")
        pdf.drawRightString(col_right["maliyet"] - 4, table_top - 12, "Maliyet")
        pdf.drawRightString(col_right["kar"] - 4, table_top - 12, "Kar")
        return table_top - 24, col_left, col_right

    page_no = 1
    y, col_left, col_right = draw_header(page_no)
    row_h = 16

    if prod_df.empty:
        pdf.setFillColor(colors.HexColor("#111827"))
        pdf.setFont(font_regular, 10)
        pdf.drawString(margin, y - 4, "Bu ay iÃ§in veri yok.")
    else:
        for i, (_, r) in enumerate(prod_df.iterrows()):
            if y < 52:
                pdf.setStrokeColor(colors.HexColor("#e5e7eb"))
                pdf.line(margin, 34, w - margin, 34)
                pdf.setFillColor(colors.HexColor("#6b7280"))
                pdf.setFont(font_regular, 8)
                pdf.drawString(margin, 22, "Eternal Fire - Finans Raporu")
                pdf.showPage()
                page_no += 1
                y, col_left, col_right = draw_header(page_no)
            if i % 2 == 0:
                pdf.setFillColor(colors.HexColor("#f9fafb"))
                pdf.rect(margin, y - row_h + 3, w - (margin * 2), row_h, stroke=0, fill=1)

            sku = str(r.get("Stok Kodu", "") or "")
            name = str(r.get("Urun", "") or "")
            if len(name) > 40:
                name = name[:37] + "..."
            adet = float(r.get("Adet", 0) or 0)
            ciro = float(r.get("Ciro", 0) or 0)
            maliyet = float(r.get("Maliyet", 0) or 0)
            kar = float(r.get("Kar", 0) or 0)
            pdf.setFillColor(colors.HexColor("#111827"))
            pdf.setFont(font_regular, 8.3)
            pdf.drawString(col_left["sku"] + 4, y - 8, sku[:20])
            pdf.drawString(col_left["urun"] + 4, y - 8, name)
            pdf.drawRightString(col_right["adet"] - 4, y - 8, format(adet, ",.0f").replace(",", "."))
            pdf.drawRightString(col_right["ciro"] - 4, y - 8, money(ciro))
            pdf.drawRightString(col_right["maliyet"] - 4, y - 8, money(maliyet))
            pdf.drawRightString(col_right["kar"] - 4, y - 8, money(kar))
            y -= row_h

    pdf.setStrokeColor(colors.HexColor("#e5e7eb"))
    pdf.line(margin, 34, w - margin, 34)
    pdf.setFillColor(colors.HexColor("#6b7280"))
    pdf.setFont(font_regular, 8)
    pdf.drawString(margin, 22, "Eternal Fire - Finans Raporu")

    pdf.save()
    return out.getvalue()


def render_header():
    now_txt = datetime.now(ZoneInfo("Europe/Istanbul")).strftime("%d.%m.%Y %H:%M")
    if APP_LOGO_URL and not APP_LOGO_URL.lower().startswith(("http://", "https://")):
        st.image(APP_LOGO_URL, width=64)
    st.markdown(
        "<h1 style='text-align:center; margin-top:2cm; margin-bottom:0;'>Eternal Fire</h1>",
        unsafe_allow_html=True,
    )
    st.markdown(
        "<p style='text-align:center; margin-top:0; color:#aeb8cb;'>Satis Operasyon ve Karlilik Kontrol Paneli</p>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<p style='text-align:center; margin-top:4px; color:#9ca3af; font-size:0.86rem;'>Son Giris: {now_txt}</p>",
        unsafe_allow_html=True,
    )

def render_login():
    st.subheader("Giris")
    with st.form("login_form", clear_on_submit=False):
        user = st.text_input("Kullanici Adi", key="login_user")
        password = st.text_input("Parola", type="password", key="login_pass")
        submitted = st.form_submit_button("Giris Yap", type="primary")
    if submitted:
        if user.strip() == APP_USER and password == APP_PASSWORD:
            st.session_state["authenticated"] = True
            st.session_state["section"] = "Veri Ekle"
            st.success("Giris basarili.")
            st.rerun()
        else:
            st.error("Kullanici adi veya parola hatali.")


inject_styles()
render_header()

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    render_login()
    st.stop()

top_c1, top_c2 = st.columns([6, 1])
with top_c2:
    if st.button("Cikis"):
        st.session_state["authenticated"] = False
        st.rerun()

sections = ["Veri Ekle", "Dashboard", "Aylik Rapor", "Bedelsiz Cikislar", "Iade Edilenler", "Urunler"]
if st.session_state.get("section") not in sections:
    st.session_state["section"] = "Veri Ekle"
section = st.radio("Bolum", sections, horizontal=True, label_visibility="collapsed", key="section")

if section == "Dashboard":
    conn = get_ready_conn()
    a1, a2 = st.columns([1, 2])
    with a1:
        if st.button("Ozeti Yenile", type="secondary"):
            with st.spinner("Ozet yenileniyor..."):
                refresh_monthly_summary_all(conn)
                st.cache_data.clear()
            st.success("Dashboard ozeti yenilendi.")
            st.rerun()
    with a2:
        if canvas is None:
            st.caption("PDF icin reportlab gerekli.")
        else:
            with st.popover("Aylik PDF Rapor"):
                month_options_df = get_available_months(conn, free_only=0)
                month_options = month_options_df["ym"].dropna().astype(str).tolist() if not month_options_df.empty else []
                if not month_options:
                    month_options = [datetime.today().strftime("%Y-%m")]
                pdf_ym = st.selectbox("Ay Sec", month_options, key="dash_pdf_ym_select")
                pdf_prod_df = get_month_products(conn, pdf_ym.strip())
                pdf_totals_df = get_month_totals(conn, pdf_ym.strip())
                pdf_order_total_df = get_month_order_total(conn, pdf_ym.strip())
                pdf_bytes = build_month_pdf(pdf_ym.strip(), pdf_totals_df, pdf_order_total_df, pdf_prod_df)
                st.download_button(
                    "PDF Indir",
                    data=pdf_bytes or b"",
                    file_name=f"eternal-fire-aylik-rapor-{pdf_ym.strip()}.pdf",
                    mime="application/pdf",
                    disabled=(pdf_bytes is None),
                )

    metrics = get_dashboard_metrics(conn)
    order_total_df = get_dashboard_order_total(conn)
    order_rows_df = get_dashboard_order_rows(conn)
    date_range_df = get_dashboard_date_range(conn)
    if metrics.empty or float(metrics.iloc[0]["total_revenue"] or 0) <= 0:
        st.info("Gosterilecek veri yok.")
    else:
        order_rows = int(order_rows_df.iloc[0]["order_rows"] or 0) if not order_rows_df.empty else 0
        q = float(metrics.iloc[0]["total_qty"] or 0)
        gross_rev = float(order_total_df.iloc[0]["gross_order_revenue"] or 0) if not order_total_df.empty else float(metrics.iloc[0]["total_revenue"] or 0)
        net_rev = float(order_total_df.iloc[0]["net_order_revenue"] or 0) if not order_total_df.empty else float(metrics.iloc[0]["total_revenue"] or 0)
        returned_rev = float(order_total_df.iloc[0]["returned_order_revenue"] or 0) if not order_total_df.empty else 0.0
        cost = float(metrics.iloc[0]["total_cost"] or 0)
        profit = net_rev - cost
        margin = (profit / net_rev * 100.0) if net_rev > 0 else 0.0
        r1c1, r1c2, r1c3 = st.columns(3)
        r1c1.metric("Toplam Satis", tr_money(gross_rev))
        r1c2.metric("Net Satis", tr_money(net_rev))
        r1c3.metric("Iade Tutar", tr_money(returned_rev))

        r2c1, r2c2, r2c3, r2c4 = st.columns(4)
        r2c1.metric("Toplam Siparis", f"{order_rows:,.0f}".replace(",", "."))
        r2c2.metric("Toplam Adet", f"{q:,.0f}".replace(",", "."))
        r2c3.metric("Toplam Maliyet", tr_money(cost))
        r2c4.metric("Net Kar", f"{tr_money(profit)} | %{margin:.1f}")
        min_d = str(date_range_df.iloc[0]["min_date"] or "") if not date_range_df.empty else ""
        max_d = str(date_range_df.iloc[0]["max_date"] or "") if not date_range_df.empty else ""
        if min_d and max_d:
            try:
                min_txt = datetime.strptime(min_d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                max_txt = datetime.strptime(max_d[:10], "%Y-%m-%d").strftime("%d.%m.%Y")
                st.caption(f"Toplam Siparis ve Toplam Adet veri araligi: {min_txt} - {max_txt}")
            except Exception:
                st.caption(f"Toplam Siparis ve Toplam Adet veri araligi: {min_d} - {max_d}")

        st.markdown("#### En Cok Satan Urunler")
        top_df = get_dashboard_top_products(conn)
        if not top_df.empty:
            top_df["Adet"] = top_df["Adet"].map(lambda v: f"{float(v):,.0f}".replace(",", "."))
            st.dataframe(top_df, use_container_width=True, hide_index=True)

        st.markdown("#### Kategori Bazli Ciro / Kar")
        cat_df = get_dashboard_categories(conn)
        if not cat_df.empty:
            cat_df["Ciro"] = cat_df["Ciro"].map(tr_money)
            cat_df["Kar"] = cat_df["Kar"].map(tr_money)
            st.dataframe(cat_df, use_container_width=True, hide_index=True)

elif section == "Veri Ekle":
    conn = get_ready_conn()
    if st.button("Tum aylik ozetleri yeniden olustur", type="secondary"):
        with st.spinner("Tum ozetler hazirlaniyor..."):
            refresh_monthly_summary_all(conn)
            st.cache_data.clear()
        st.success("Tum aylik ozetler yenilendi.")
        st.rerun()

    st.markdown("#### Veri Yonetimi")
    m1, m2, m3 = st.columns(3)
    with m1:
        del_ym = st.text_input("Toplu silme ayi (YYYY-MM)", value=datetime.today().strftime("%Y-%m"), key="del_ym")
        if st.button("Secili Ayi Sil"):
            if not is_valid_ym(del_ym):
                st.warning("Ay formati gecersiz. Ornek: 2026-02")
            else:
                conn.execute("DELETE FROM sales WHERE ym = ?", (del_ym.strip(),))
                conn.commit()
                refresh_monthly_summary_for_month(conn, del_ym.strip())
                st.cache_data.clear()
                st.success(f"{del_ym.strip()} ayi verileri silindi.")
                st.rerun()
    with m2:
        recent_df = get_recent_sales(conn, 300)
        if recent_df.empty:
            st.caption("Tek satir silme icin veri yok.")
        else:
            recent_df["label"] = recent_df.apply(
                lambda r: f"#{int(r['id'])} | {r['order_date']} | {r['sku']} | {float(r['qty']):,.0f} adet".replace(",", "."),
                axis=1,
            )
            sel_label = st.selectbox("Tek satir sec", recent_df["label"].tolist(), key="single_delete_row")
            sel_row = recent_df[recent_df["label"] == sel_label].iloc[0]
            if st.button("Secili Satiri Sil"):
                row_id = int(sel_row["id"])
                ym = str(sel_row["ym"]) if "ym" in sel_row and str(sel_row["ym"]) != "nan" else str(sel_row["order_date"])[:7]
                conn.execute("DELETE FROM sales WHERE id=?", (row_id,))
                conn.commit()
                if is_valid_ym(ym):
                    refresh_monthly_summary_for_month(conn, ym)
                st.cache_data.clear()
                st.success(f"Satir silindi: #{row_id}")
                st.rerun()
    with m3:
        wipe_ok = st.checkbox("Tum veriyi silmeyi onayliyorum", key="wipe_all_confirm")
        if st.button("Tum Verileri Sil", type="secondary", disabled=not wipe_ok):
            conn.execute("DELETE FROM sales")
            conn.execute("DELETE FROM sales_monthly_sku")
            conn.execute("DELETE FROM order_registry")
            conn.execute("DELETE FROM app_metrics")
            conn.commit()
            st.cache_data.clear()
            st.success("Tum satis verileri silindi.")
            st.rerun()

    uploads = get_uploads(conn)
    if not uploads.empty:
        last_week = str(uploads.iloc[0]["week_label"])
        last_file = str(uploads.iloc[0]["source_file"])
        last_hash = str(uploads.iloc[0].get("source_hash", "") or "").strip()
        last_rows = int(uploads.iloc[0]["row_count"])
        st.caption(f"Son yukleme: {last_week} | {last_file} | {last_rows} satir")

        if st.button("Son yuklemeyi sil", type="secondary"):
            with st.spinner("Siliniyor..."):
                if not last_hash:
                    st.warning("Son yukleme bilgisi bulunamadi.")
                else:
                    src_hash = last_hash
                    months_df = df_query(
                        conn,
                        "SELECT DISTINCT ym FROM sales WHERE source_hash=?",
                        (src_hash,),
                    )
                    conn.execute("DELETE FROM sales WHERE source_hash=?", (src_hash,))
                    conn.execute("DELETE FROM order_registry WHERE source_hash=?", (src_hash,))
                    conn.execute("DELETE FROM app_metrics WHERE source_hash=?", (src_hash,))
                    conn.commit()
                    for m in months_df["ym"].dropna().astype(str).tolist():
                        refresh_monthly_summary_for_month(conn, m)
                    st.cache_data.clear()
                    st.success(f"Son yukleme silindi: {last_file} ({last_week})")
            st.rerun()

    auto_week_label = f"{datetime.today().isocalendar().year}-W{datetime.today().isocalendar().week:02d}"
    uploaded = st.file_uploader("Excel sec", type=["xlsx"])
    if st.button("Yukle ve Isle", type="primary", disabled=uploaded is None):
        with st.spinner("Excel isleniyor..."):
            bytes_ = uploaded.getvalue()
            source_hash = hashlib.sha256(bytes_).hexdigest()
            source_file = uploaded.name
            parsed = parse_uploaded_excel(bytes_, auto_week_label, datetime.today().date().isoformat())
            order_registry_rows = parsed.attrs.get("order_registry_rows", [])
            summary_metrics = parsed.attrs.get("summary_metrics", {})
            upsert_order_registry(conn, order_registry_rows, source_file, source_hash)
            upsert_app_metrics(conn, summary_metrics, source_file, source_hash)
            if parsed.empty:
                order_registry_count = len({str(o.get("order_no", "") or "").strip() for o in order_registry_rows if str(o.get("order_no", "") or "").strip()})
                if summary_metrics:
                    st.success(
                        f"Ikas ozet metrikleri guncellendi. Iade Tutar: {tr_money(summary_metrics.get('ikas_summary_refund_amount', 0))}"
                    )
                    st.cache_data.clear()
                elif order_registry_count:
                    st.success(f"Ikas siparis sayimi guncellendi: {order_registry_count}")
                    st.cache_data.clear()
                else:
                    st.error("Islenecek satir bulunamadi.")
            else:
                parsed["source_file"] = source_file
                parsed["source_hash"] = source_hash
                parsed["order_item_key"] = parsed.apply(
                    lambda r: build_order_item_key(
                        str(r.get("order_no", "") or ""),
                        str(r.get("order_date", "") or ""),
                        str(r.get("sku", "") or ""),
                        float(r.get("qty", 0) or 0),
                        float(r.get("unit_price", 0) or 0),
                        str(r.get("customer_email", "") or ""),
                        source_hash,
                        int(r.get("excel_row_no", 0) or 0),
                    ),
                    axis=1,
                )
                rows = [
                    (
                        r["week_label"],
                        r["order_date"],
                        r["ym"],
                        r["order_no"],
                        r["order_item_key"],
                        r["customer_email"],
                        int(r["is_free_exit"]),
                        int(r.get("is_returned", 0) or 0),
                        r["sku"],
                        r["product_name"],
                        float(r["qty"]),
                        float(r["unit_price"]),
                        float(r["revenue"]),
                        float(r.get("order_total", 0) or 0),
                        r["source_file"],
                        r["source_hash"],
                    )
                    for _, r in parsed.iterrows()
                ]
                inserted_months = sorted({str(x).strip() for x in parsed["ym"].dropna().astype(str).tolist() if is_valid_ym(str(x).strip())})
                affected_months = set(inserted_months)

                # One-shot cleanup for previously mis-dated imports:
                # remove rows with same order numbers before re-inserting corrected data.
                order_nos = (
                    parsed["order_no"]
                    .fillna("")
                    .astype(str)
                    .str.strip()
                )
                order_nos = sorted({x for x in order_nos.tolist() if x})
                if order_nos:
                    existing = select_months_for_order_nos(conn, order_nos)
                    affected_months.update([str(m).strip() for m in existing["ym"].dropna().astype(str).tolist() if is_valid_ym(str(m).strip())])
                    delete_sales_by_order_nos(conn, order_nos)

                # Also replace old rows from same exact file hash.
                source_months = df_query(conn, "SELECT DISTINCT ym FROM sales WHERE source_hash=?", (source_hash,))
                affected_months.update([str(m).strip() for m in source_months["ym"].dropna().astype(str).tolist() if is_valid_ym(str(m).strip())])
                conn.execute("DELETE FROM sales WHERE source_hash=?", (source_hash,))
                conn.commit()
                before_count = int(
                    df_query(conn, "SELECT COUNT(*) AS n FROM sales WHERE source_hash=?", (source_hash,)).iloc[0]["n"]
                )
                conn.executemany(
                    """
                    INSERT INTO sales(week_label, order_date, ym, order_no, order_item_key, customer_email, is_free_exit, is_returned, sku, product_name, qty, unit_price, revenue, order_total, source_file, source_hash)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(order_item_key) DO UPDATE SET
                        week_label=excluded.week_label,
                        order_date=excluded.order_date,
                        ym=excluded.ym,
                        order_no=excluded.order_no,
                        customer_email=excluded.customer_email,
                        is_free_exit=excluded.is_free_exit,
                        is_returned=excluded.is_returned,
                        sku=excluded.sku,
                        product_name=excluded.product_name,
                        qty=excluded.qty,
                        unit_price=excluded.unit_price,
                        revenue=excluded.revenue,
                        order_total=excluded.order_total,
                        source_file=excluded.source_file,
                        source_hash=excluded.source_hash
                    """,
                    rows,
                )
                conn.commit()
                upsert_products_from_rows(conn, parsed[["sku", "product_name"]])
                for m in sorted(affected_months):
                    if is_valid_ym(m):
                        refresh_monthly_summary_for_month(conn, m)
                st.cache_data.clear()
                free_rows = int(parsed["is_free_exit"].sum())
                returned_rows = int(parsed["is_returned"].sum()) if "is_returned" in parsed.columns else 0
                order_registry_count = len({str(o.get("order_no", "") or "").strip() for o in order_registry_rows if str(o.get("order_no", "") or "").strip()})
                after_count = int(
                    df_query(conn, "SELECT COUNT(*) AS n FROM sales WHERE source_hash=?", (source_hash,)).iloc[0]["n"]
                )
                inserted_rows = max(0, after_count - before_count)
                skipped_rows = max(0, len(rows) - inserted_rows)
                if inserted_rows == 0:
                    st.warning(
                        f"Yeni kayit eklenmedi. Ikas siparis: {order_registry_count} | Atlanan tekrar: {skipped_rows} | Bedelsiz cikis: {free_rows} | Iade: {returned_rows}"
                    )
                else:
                    st.success(
                        f"Yukleme tamamlandi. Ikas siparis: {order_registry_count} | Eklenen: {inserted_rows} | Atlanan tekrar: {skipped_rows} | Bedelsiz cikis: {free_rows} | Iade: {returned_rows}"
                    )

elif section == "Aylik Rapor":
    conn = get_ready_conn()
    month_options_df = get_available_months(conn, free_only=0)
    month_options = month_options_df["ym"].dropna().astype(str).tolist() if not month_options_df.empty else []
    if not month_options:
        month_options = [datetime.today().strftime("%Y-%m")]
    ym = st.selectbox("Ay Sec", month_options, key="month_ym_select")

    prod_df = get_month_products(conn, ym.strip())
    totals_df = get_month_totals(conn, ym.strip())
    month_order_total_df = get_month_order_total(conn, ym.strip())

    if prod_df.empty:
        st.info("Bu ay icin veri yok.")
    else:
        t_qty = float(totals_df.iloc[0]["total_qty"] or 0)
        t_rev = float(month_order_total_df.iloc[0]["total_order_revenue"] or 0) if not month_order_total_df.empty else float(totals_df.iloc[0]["total_revenue"] or 0)
        t_cost = float(totals_df.iloc[0]["total_cost"] or 0)
        t_profit = t_rev - t_cost
        x1, x2, x3, x4 = st.columns(4)
        x1.metric("Toplam Adet", f"{t_qty:,.0f}".replace(",", "."))
        x2.metric("Toplam Ciro", tr_money(t_rev))
        x3.metric("Toplam Maliyet", tr_money(t_cost))
        x4.metric("Gercek Kar", tr_money(t_profit))

        disp = prod_df.copy()
        disp["Adet"] = disp["Adet"].map(lambda v: f"{float(v):,.0f}".replace(",", "."))
        disp["Ciro"] = disp["Ciro"].map(tr_money)
        disp["Maliyet"] = disp["Maliyet"].map(tr_money)
        disp["Kar"] = disp["Kar"].map(tr_money)
        st.dataframe(disp, use_container_width=True, hide_index=True)

        sku_opts = prod_df["Stok Kodu"].astype(str).tolist()
        sku_sel = st.selectbox("SKU detay sec", sku_opts)
        sku_df = get_month_sku_details(conn, ym.strip(), sku_sel)
        if not sku_df.empty:
            sku_df["Adet"] = sku_df["Adet"].map(lambda v: f"{float(v):,.0f}".replace(",", "."))
            sku_df["Birim Fiyat"] = sku_df["Birim Fiyat"].map(tr_money)
            sku_df["Ciro"] = sku_df["Ciro"].map(tr_money)
            st.dataframe(sku_df, use_container_width=True, hide_index=True)

elif section == "Bedelsiz Cikislar":
    conn = get_ready_conn()
    free_months_df = get_available_months(conn, free_only=1)
    free_months = free_months_df["ym"].dropna().astype(str).tolist() if not free_months_df.empty else []
    if not free_months:
        free_months = [datetime.today().strftime("%Y-%m")]
    ym = st.selectbox("Ay Sec", free_months, key="free_ym_select")

    free_df = get_free_exit_manage_rows(conn, ym.strip())
    if free_df.empty:
        st.info("Bu ay bedelsiz cikis yok.")
    else:
        fq = float(free_df["Adet"].sum())
        fr = float(free_df["_BedelsizTutar"].sum())
        c1, c2 = st.columns(2)
        c1.metric("Bedelsiz Toplam Adet", f"{fq:,.0f}".replace(",", "."))
        c2.metric("Bedelsiz Toplam Tutar", tr_money(fr))

        editor_df = free_df.copy()
        editor_df["Adet"] = editor_df["Adet"].map(float)
        editor_df["Birim Fiyat"] = editor_df["Birim Fiyat"].map(float)
        editor_df["_BedelsizTutar"] = editor_df["_BedelsizTutar"].map(float)
        editor_df = editor_df[
            ["ID", "Tarih", "E-Posta", "Stok Kodu", "Urun", "Adet", "Birim Fiyat", "Aciklama"]
        ]
        edited = st.data_editor(
            editor_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "ID": st.column_config.NumberColumn("ID", disabled=True),
                "Tarih": st.column_config.TextColumn("Tarih", disabled=True),
                "E-Posta": st.column_config.TextColumn("E-Posta", disabled=True),
                "Stok Kodu": st.column_config.TextColumn("Stok Kodu", disabled=True),
                "Urun": st.column_config.TextColumn("Urun", disabled=True),
                "Adet": st.column_config.NumberColumn("Adet", disabled=True),
                "Birim Fiyat": st.column_config.NumberColumn("Birim Fiyat", disabled=True),
                "Aciklama": st.column_config.TextColumn("Kime / Aciklama"),
            },
            num_rows="fixed",
            key="free_exit_editor",
        )
        if st.button("Bedelsiz Aciklamalari Kaydet", type="primary"):
            rows = []
            for _, r in edited.iterrows():
                rows.append((str(r.get("Aciklama", "") or ""), int(r["ID"])))
            conn.executemany("UPDATE sales SET free_exit_note=? WHERE id=?", rows)
            conn.commit()
            st.cache_data.clear()
            st.success("Bedelsiz aciklamalari kaydedildi.")
            st.rerun()

elif section == "Iade Edilenler":
    conn = get_ready_conn()
    returned_months_df = get_returned_months(conn)
    returned_months = returned_months_df["ym"].dropna().astype(str).tolist() if not returned_months_df.empty else []
    if not returned_months:
        returned_months = [datetime.today().strftime("%Y-%m")]
    ym = st.selectbox("Ay Sec", returned_months, key="returned_ym_select")

    returned_df = get_returned_rows(conn, ym.strip())
    if returned_df.empty:
        st.info("Bu ay iade edilen siparis yok.")
    else:
        rq = float(returned_df["Adet"].sum())
        rr = float(returned_df["Iade Tutar"].sum())
        c1, c2 = st.columns(2)
        c1.metric("Iade Toplam Adet", f"{rq:,.0f}".replace(",", "."))
        c2.metric("Iade Toplam Tutar", tr_money(rr))

        show_df = returned_df.copy()
        show_df["Adet"] = show_df["Adet"].map(lambda v: f"{float(v):,.0f}".replace(",", "."))
        show_df["Birim Fiyat"] = show_df["Birim Fiyat"].map(tr_money)
        show_df["Iade Tutar"] = show_df["Iade Tutar"].map(tr_money)
        st.dataframe(show_df, use_container_width=True, hide_index=True)

else:
    conn = get_ready_conn()
    st.subheader("Urunler")
    products = get_products_master(conn)
    if products.empty:
        st.info("Once Excel yukleyin.")
    else:
        edit_df = products.copy()
        edit_df["category"] = edit_df["category"].fillna("Genel")
        edit_df["unit_cost"] = edit_df["unit_cost"].fillna(0.0)
        edit_df["active"] = edit_df["active"].fillna(1).astype(int).map(lambda x: x == 1)
        edited = st.data_editor(
            edit_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "sku": st.column_config.TextColumn("Stok Kodu", disabled=True),
                "product_name": st.column_config.TextColumn("Urun", disabled=True),
                "category": st.column_config.TextColumn("Kategori"),
                "unit_cost": st.column_config.NumberColumn("Birim Maliyet", min_value=0.0, step=0.01),
                "active": st.column_config.CheckboxColumn("Aktif"),
            },
            num_rows="fixed",
        )
        if st.button("Urun Maliyet Kaydet", type="primary"):
            now = datetime.now().isoformat(timespec="seconds")
            rows = []
            cost_rows = []
            for _, r in edited.iterrows():
                rows.append(
                    (
                        r["sku"],
                        r["product_name"],
                        str(r.get("category", "Genel") or "Genel"),
                        float(r["unit_cost"]),
                        1 if bool(r.get("active", True)) else 0,
                        now,
                    )
                )
                cost_rows.append((r["sku"], float(r["unit_cost"])))
            conn.executemany(
                """
                INSERT INTO products(sku, product_name, category, unit_cost, active, updated_at)
                VALUES(?,?,?,?,?,?)
                ON CONFLICT(sku) DO UPDATE SET
                    product_name=excluded.product_name,
                    category=excluded.category,
                    unit_cost=excluded.unit_cost,
                    active=excluded.active,
                    updated_at=excluded.updated_at
                """,
                rows,
            )
            conn.commit()
            upsert_product_costs(conn, cost_rows)
            st.cache_data.clear()
            st.success("Urun master kaydedildi.")





